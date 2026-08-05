"""
Shared data processing engines and chart theme for OSA & SOS.
Imported by all pages.
"""

import io
import os
import re
import json
import pathlib
import tempfile
import itertools
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
#  CHART THEME
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
#  THEME  (dark default, light mode toggled via session_state)
# ─────────────────────────────────────────────────────────────────────────────

THEME_DARK = {
    "bg": "#0f1117", "surface": "#1a1d27", "border": "#2a2d3a",
    "text": "#e8eaf0", "muted": "#6b7280",
    "green": "#34c97b", "amber": "#f59e0b", "red": "#f87171",
    "blue": "#4f8ef7", "purple": "#a78bfa", "cyan": "#22d3ee",
}
THEME_LIGHT = {
    "bg": "#f0f4ff", "surface": "#ffffff", "border": "#c7d7f5",
    "text": "#0f1f4b", "muted": "#5a6a8a",
    "green": "#16a34a", "amber": "#d97706", "red": "#dc2626",
    "blue": "#1d4ed8", "purple": "#7c3aed", "cyan": "#0891b2",
}

# Default chart theme (dark) — pages call get_plotly_layout() at render time
THEME = THEME_DARK

def get_theme():
    try:
        import streamlit as st
        if st.session_state.get("light_mode", False):
            return THEME_LIGHT
        return THEME_DARK
    except Exception:
        return THEME_DARK

def get_plotly_layout(height=380, title=""):
    t = get_theme()
    layout = dict(
        paper_bgcolor=t["surface"],
        plot_bgcolor=t["surface"],
        height=height,
        font=dict(family="DM Sans, sans-serif", color=t["text"], size=12),
        margin=dict(l=20, r=20, t=40 if title else 20, b=20),
        legend=dict(
            bgcolor=t["bg"],
            bordercolor=t["border"],
            borderwidth=1,
            font=dict(size=11, color=t["text"]),   # explicit text color
            title_font=dict(color=t["text"]),
        ),
        xaxis=dict(
            gridcolor=t["border"],
            linecolor=t["border"],
            tickfont=dict(size=11, color=t["text"]),
            title_font=dict(color=t["text"]),
        ),
        yaxis=dict(
            gridcolor=t["border"],
            linecolor=t["border"],
            tickfont=dict(size=11, color=t["text"]),
            title_font=dict(color=t["text"]),
        ),
    )
    # Only add title key if there's actually a title — avoids "undefined" rendering
    if title:
        layout["title"] = dict(
            text=title,
            font=dict(size=13, color=t["muted"]),
            x=0, pad=dict(l=0)
        )
    return layout

def osa_color(val):
    t = get_theme()
    if __import__('pandas').isna(val): return t["muted"]
    if val >= 95: return t["green"]
    if val >= 80: return t["amber"]
    return t["red"]

# Keep for backward compat
PLOTLY_LAYOUT = dict(
    paper_bgcolor=THEME_DARK["surface"], plot_bgcolor=THEME_DARK["surface"],
    font=dict(family="DM Sans, sans-serif", color=THEME_DARK["text"], size=12),
    margin=dict(l=20, r=20, t=40, b=20),
    legend=dict(bgcolor=THEME_DARK["bg"], bordercolor=THEME_DARK["border"], borderwidth=1, font=dict(size=11)),
    xaxis=dict(gridcolor=THEME_DARK["border"], linecolor=THEME_DARK["border"], tickfont=dict(size=11)),
    yaxis=dict(gridcolor=THEME_DARK["border"], linecolor=THEME_DARK["border"], tickfont=dict(size=11)),
)

# ─────────────────────────────────────────────────────────────────────────────
#  GLOBAL CSS  — uses CSS variables so one class on <body> flips the theme
# ─────────────────────────────────────────────────────────────────────────────

GLOBAL_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

/* ── CSS variables: dark (default) ── */
:root {
  --bg:      #0f1117;
  --surface: #1a1d27;
  --border:  #2a2d3a;
  --text:    #e8eaf0;
  --muted:   #6b7280;
  --blue:    #4f8ef7;
  --blue2:   #6b9ff8;
  --green:   #34c97b;
  --amber:   #f59e0b;
  --red:     #f87171;
  --sidebar-bg: #1a1d27;
}

/* ── CSS variables: light mode ── */
body.light-mode {
  --bg:      #f0f4ff;
  --surface: #ffffff;
  --border:  #c7d7f5;
  --text:    #0f1f4b;
  --muted:   #5a6a8a;
  --blue:    #1d4ed8;
  --blue2:   #2563eb;
  --green:   #16a34a;
  --amber:   #d97706;
  --red:     #dc2626;
  --sidebar-bg: #e8eeff;
}

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }
#MainMenu, footer, header { visibility: hidden; }

/* ── app background ── */
.stApp { background: var(--bg) !important; }
.stApp > div { background: var(--bg) !important; }
.block-container { padding-top: 1.5rem; padding-bottom: 3rem; max-width: 1100px; }

/* ── sidebar ── */
section[data-testid="stSidebar"] { background: var(--sidebar-bg) !important; }
section[data-testid="stSidebar"] * { color: var(--text) !important; }
section[data-testid="stSidebar"] a { color: var(--blue) !important; }

/* ── main text ── */
.stMarkdown, .stMarkdown p, div[data-testid="stMarkdownContainer"] p { color: var(--text) !important; }

/* ── page title ── */
.pg-title { margin-bottom: 1.5rem; }
.pg-title .eyebrow { font-family:'DM Mono',monospace; font-size:10px; letter-spacing:3px; color:var(--blue); text-transform:uppercase; margin-bottom:4px; }
.pg-title h1 { font-size:26px; font-weight:600; color:var(--text); margin:0 0 4px 0; }
.pg-title p  { font-size:13px; color:var(--muted); margin:0; }

/* ── metric cards ── */
.metric-row { display:flex; gap:12px; margin-bottom:20px; flex-wrap:wrap; }
.metric-card { flex:1; min-width:130px; background:var(--surface); border:1px solid var(--border); border-radius:12px; padding:16px 18px; }
.metric-card .m-label { font-family:'DM Mono',monospace; font-size:9px; letter-spacing:2px; text-transform:uppercase; color:var(--muted); margin-bottom:6px; }
.metric-card .m-value { font-size:28px; font-weight:600; color:var(--text); line-height:1; }
.metric-card .m-sub   { font-size:11px; color:var(--muted); margin-top:4px; }

/* ── section label ── */
.sec-label { font-family:'DM Mono',monospace; font-size:10px; letter-spacing:2px; text-transform:uppercase; color:var(--muted); margin:20px 0 8px 0; }

/* ── file uploader ── */
div[data-testid="stFileUploader"] { border:2px dashed var(--border) !important; border-radius:12px; background:var(--surface); padding:12px; }
div[data-testid="stFileUploader"]:hover { border-color:var(--blue) !important; }
div[data-testid="stFileUploader"] label { color:var(--text) !important; }
div[data-testid="stFileUploader"] small { color:var(--muted) !important; }

/* ── primary button ── */
div[data-testid="stButton"] > button { background:var(--blue) !important; color:#fff !important; border:none !important; border-radius:10px !important; padding:12px 20px !important; font-size:14px !important; font-weight:600 !important; font-family:'DM Sans',sans-serif !important; width:100%; }
div[data-testid="stButton"] > button:hover { background:var(--blue2) !important; }

/* ── download button ── */
div[data-testid="stDownloadButton"] > button { background:var(--green) !important; color:#fff !important; border:none !important; border-radius:10px !important; padding:12px 20px !important; font-size:14px !important; font-weight:600 !important; width:100%; }

/* ── select / multiselect ── */
div[data-testid="stSelectbox"] > div, div[data-testid="stMultiSelect"] > div { background:var(--surface) !important; border-color:var(--border) !important; color:var(--text) !important; }

/* ── tabs ── */
button[data-baseweb="tab"] { color:var(--muted) !important; font-size:13px !important; }
button[data-baseweb="tab"][aria-selected="true"] { color:var(--blue) !important; }
div[data-baseweb="tab-highlight"] { background-color:var(--blue) !important; }
div[data-baseweb="tab-border"]    { background-color:var(--border) !important; }

/* ── alerts ── */
.result-box { border-radius:10px; padding:12px 16px; font-size:13px; margin:8px 0; display:flex; gap:10px; align-items:flex-start; }
.result-box.success { background:rgba(52,201,123,.12); border:1px solid rgba(52,201,123,.4); color:var(--green); }
.result-box.error   { background:rgba(220,38,38,.08);  border:1px solid rgba(220,38,38,.3);  color:var(--red); }
.result-box.info    { background:rgba(29,78,216,.08);  border:1px solid rgba(29,78,216,.3);  color:var(--blue); }

hr { border-color:var(--border) !important; margin:20px 0 !important; }
.footer { text-align:center; font-size:11px; color:var(--muted); margin-top:12px; }

/* ── multiselect tags ── */
span[data-baseweb="tag"] { background:var(--blue) !important; }

/* ── light mode: extra overrides ── */
body.light-mode .stApp,
body.light-mode .stApp > div { background: var(--bg) !important; }
body.light-mode section[data-testid="stSidebar"] { background: var(--sidebar-bg) !important; box-shadow: 2px 0 12px rgba(29,78,216,.08); }
body.light-mode .metric-card { box-shadow: 0 1px 4px rgba(29,78,216,.08); }
body.light-mode div[data-testid="stTextInput"] input,
body.light-mode div[data-testid="stSelectbox"] > div,
body.light-mode div[data-testid="stMultiSelect"] > div { background: #f8faff !important; color: var(--text) !important; border-color: var(--border) !important; }
body.light-mode hr { border-color: var(--border) !important; }

/* ── light mode: force ALL text dark ── */
body.light-mode p,
body.light-mode h1, body.light-mode h2, body.light-mode h3,
body.light-mode label, body.light-mode span,
body.light-mode div[data-testid="stMarkdownContainer"],
body.light-mode div[data-testid="stMarkdownContainer"] * { color: var(--text) !important; }

/* ── light mode: nav card buttons — dark text on white/blue ── */
body.light-mode section[data-testid="column"] div[data-testid="stButton"] > button {
    background: #ffffff !important;
    border: 2px solid var(--blue) !important;
    color: var(--text) !important;
}
body.light-mode section[data-testid="column"] div[data-testid="stButton"] > button:hover {
    background: var(--blue) !important;
    color: #ffffff !important;
}

/* ── light mode: primary action buttons stay blue with white text ── */
body.light-mode div[data-testid="stButton"] > button {
    color: #ffffff !important;
}
/* but nav card buttons override to dark text (above rule is more specific) */

/* ── light mode: legend pills ── */
body.light-mode .legend-pill { background: #ffffff; border-color: var(--border); color: var(--text) !important; }

/* ── light mode: home title ── */
body.light-mode .home-hero h1 { color: var(--text) !important; }
body.light-mode .home-hero p  { color: var(--muted) !important; }
body.light-mode .home-hero .eyebrow { color: var(--blue) !important; }

/* ── light mode: pg-title ── */
body.light-mode .pg-title h1 { color: var(--text) !important; }
body.light-mode .pg-title p  { color: var(--muted) !important; }

/* ── light mode: metric cards ── */
body.light-mode .metric-card .m-value { color: var(--text) !important; }
body.light-mode .metric-card .m-label { color: var(--muted) !important; }
body.light-mode .metric-card .m-sub   { color: var(--muted) !important; }

/* ── light mode: result boxes ── */
body.light-mode .result-box.info    { background: rgba(29,78,216,.06) !important; color: #1d4ed8 !important; }
body.light-mode .result-box.success { background: rgba(22,163,74,.06) !important; color: #15803d !important; }
body.light-mode .result-box.error   { background: rgba(220,38,38,.06) !important; color: #b91c1c !important; }
</style>
"""



# ── Sidebar toggle injected via components.v1.html (bypasses Streamlit's HTML sanitiser)
SIDEBAR_TOGGLE_JS = """
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  html, body { background: transparent; overflow: hidden; }
  #btn {
    position: fixed; top: 0; left: 0;
    width: 38px; height: 38px;
    background: #1a1d27;
    border: 1px solid #2a2d3a;
    border-radius: 8px; cursor: pointer;
    display: flex; flex-direction: column;
    align-items: center; justify-content: center; gap: 5px;
    transition: border-color .2s, background .2s;
  }
  #btn:hover { border-color: #4f8ef7; background: rgba(79,142,247,.15); }
  .bar {
    display: block; width: 16px; height: 2px;
    background: #e8eaf0; border-radius: 2px;
    transition: transform .25s, opacity .25s;
  }
  #btn.closed .b1 { transform: rotate(45deg) translate(5px, 5px); }
  #btn.closed .b2 { opacity: 0; }
  #btn.closed .b3 { transform: rotate(-45deg) translate(5px, -5px); }
</style>

<div id="btn" title="Toggle sidebar">
  <span class="bar b1"></span>
  <span class="bar b2"></span>
  <span class="bar b3"></span>
</div>

<script>
(function() {
  var btn = document.getElementById('btn');

  // Click Streamlit's own internal collapse button
  function clickStreamlitToggle() {
    var p = window.parent.document;
    var collapseBtn = p.querySelector('[data-testid="stSidebarCollapseButton"] button');
    if (!collapseBtn) {
        collapseBtn = p.querySelector('button[data-testid="stSidebarCollapseButton"]');
    }
    if (collapseBtn) {
      collapseBtn.click();
      return true;
    }
    return false;
  }

  function isSidebarCollapsed() {
    var p = window.parent.document;
    var sidebar = p.querySelector('section[data-testid="stSidebar"]');
    if (!sidebar) return false;
    // Streamlit adds data-collapsed or aria-expanded, or we check transform
    var style = window.parent.getComputedStyle(sidebar);
    var rect = sidebar.getBoundingClientRect();
    return rect.right <= 10 || sidebar.getAttribute('aria-expanded') === 'false';
  }

  function syncIcon() {
    if (isSidebarCollapsed()) {
      btn.classList.add('closed');
    } else {
      btn.classList.remove('closed');
    }
  }

  btn.addEventListener('click', function() {
    var didClick = clickStreamlitToggle();
    // Update icon after short delay to let Streamlit animate
    setTimeout(syncIcon, 350);
  });

  // Keep icon in sync with actual sidebar state
  function watchSidebar() {
    var p = window.parent.document;
    var sidebar = p.querySelector('section[data-testid="stSidebar"]');
    if (!sidebar) {
      setTimeout(watchSidebar, 300);
      return;
    }
    syncIcon();
    // Watch for attribute/style changes on the sidebar
    var observer = new MutationObserver(function() { syncIcon(); });
    observer.observe(sidebar, { attributes: true, attributeFilter: ['style', 'class', 'aria-expanded'] });
    // Also watch the parent for transform changes
    observer.observe(sidebar.parentElement || p.body, { childList: true, subtree: false });
  }

  // Wait for parent page to fully render
  setTimeout(watchSidebar, 500);
})();
</script>
"""

def inject_sidebar_toggle():
    """
    Injects a toggle button OUTSIDE the sidebar (in the main area, top-left),
    so it stays visible even when the sidebar is collapsed.
    Uses components.html with JS that targets Streamlit's collapse button.
    The iframe is hidden by collapsing the Streamlit component container via JS.
    """
    import streamlit as st
    import streamlit.components.v1 as components

    components.html("""
<!DOCTYPE html>
<html>
<head>
<style>
  * { margin:0; padding:0; box-sizing:border-box; }
  html, body { background:transparent; overflow:hidden; width:42px; height:42px; }
  #btn {
    width:42px; height:42px;
    background:#1a1d27; border:1.5px solid #2a2d3a; border-radius:8px;
    cursor:pointer; display:flex; flex-direction:column;
    align-items:center; justify-content:center; gap:5px;
    transition:border-color .2s, background .2s;
  }
  #btn:hover { border-color:#4f8ef7; background:rgba(79,142,247,.15); }
  .bar { display:block; width:16px; height:2px; background:#e8eaf0; border-radius:2px; transition:transform .25s, opacity .25s; }
  #btn.closed .b1 { transform:rotate(45deg) translate(5px,5px); }
  #btn.closed .b2 { opacity:0; }
  #btn.closed .b3 { transform:rotate(-45deg) translate(5px,-5px); }
</style>
</head>
<body>
<div id="btn"><span class="bar b1"></span><span class="bar b2"></span><span class="bar b3"></span></div>
<script>
(function(){
  var btn = document.getElementById('btn');

  function getCollapseBtn() {
    var p = window.parent.document;
    var b = p.querySelector('[data-testid="stSidebarCollapseButton"] button');
    if (!b) b = p.querySelector('button[data-testid="stSidebarCollapseButton"]');
    return b;
  }

  function isSidebarCollapsed() {
    var p = window.parent.document;
    var sidebar = p.querySelector('section[data-testid="stSidebar"]');
    if (!sidebar) return false;
    var rect = sidebar.getBoundingClientRect();
    return rect.right < 20;
  }

  function syncIcon() {
    if (isSidebarCollapsed()) btn.classList.add('closed');
    else btn.classList.remove('closed');
  }

  btn.addEventListener('click', function(){
    var cb = getCollapseBtn();
    if (cb) {
      cb.click();
      setTimeout(syncIcon, 400);
    }
  });

  // Also: resize this iframe container to 0 height in parent, then position fixed
  function positionSelf() {
    try {
      var p = window.parent.document;
      // Find our iframe in parent
      var frames = p.querySelectorAll('iframe');
      var me = null;
      for (var i=0; i<frames.length; i++) {
        try {
          if (frames[i].contentWindow === window) { me = frames[i]; break; }
        } catch(e) {}
      }
      if (me) {
        // Position the iframe fixed top-left
        me.style.cssText = 'position:fixed!important;top:10px!important;left:10px!important;width:42px!important;height:42px!important;border:none!important;z-index:99999!important;background:transparent!important;';
        // Collapse the wrapper div height to 0
        var wrapper = me.parentElement;
        while (wrapper && wrapper !== p.body) {
          if (wrapper.style !== undefined) {
            var h = wrapper.getBoundingClientRect().height;
            if (h > 42 && h < 200) {
              wrapper.style.height = '0';
              wrapper.style.overflow = 'visible';
              wrapper.style.marginTop = '0';
              wrapper.style.marginBottom = '0';
              wrapper.style.paddingTop = '0';
              wrapper.style.paddingBottom = '0';
            }
          }
          wrapper = wrapper.parentElement;
        }
      }
    } catch(e) {}
  }

  function init() {
    positionSelf();
    syncIcon();
    // Watch sidebar for collapse/expand
    var p = window.parent.document;
    var sidebar = p.querySelector('section[data-testid="stSidebar"]');
    if (sidebar) {
      var obs = new MutationObserver(syncIcon);
      obs.observe(sidebar, {attributes:true, attributeFilter:['style','class']});
      obs.observe(sidebar.parentElement || p.body, {childList:true, subtree:true, attributeFilter:['style']});
    }
  }

  if (document.readyState === 'loading') document.addEventListener('DOMContentLoaded', function(){ setTimeout(init,300); });
  else setTimeout(init, 300);
  setTimeout(init, 800);
  setTimeout(init, 1500);
})();
</script>
</body>
</html>
""", height=50, scrolling=False)


def inject_theme_toggle():
    """
    Renders a 🌙/☀️ toggle in the sidebar that flips between dark and light mode.
    Applies/removes the 'light-mode' class on <body> via JS, which triggers all
    CSS variable overrides in GLOBAL_CSS. Also re-runs the page so Plotly charts
    re-render with the correct colours.
    """
    import streamlit as st
    import streamlit.components.v1 as components

    # Initialise session state
    if "light_mode" not in st.session_state:
        st.session_state["light_mode"] = False

    is_light = st.session_state["light_mode"]
    label    = "☀️  Light mode" if not is_light else "🌙  Dark mode"
    tooltip  = "Switch to light mode" if not is_light else "Switch to dark mode"

    with st.sidebar:
        if st.button(label, key="_theme_toggle", help=tooltip, use_container_width=True):
            st.session_state["light_mode"] = not st.session_state["light_mode"]
            st.rerun()

    # Apply the CSS class to <body> in the parent page
    mode_class = "light-mode" if st.session_state["light_mode"] else ""
    components.html(f"""
<script>
(function() {{
  function apply() {{
    var body = window.parent.document.body;
    if (!body) return;
    if ("{mode_class}" === "light-mode") {{
      body.classList.add("light-mode");
    }} else {{
      body.classList.remove("light-mode");
    }}
  }}
  apply();
  setTimeout(apply, 200);
  setTimeout(apply, 600);
}})();
</script>
""", height=0, scrolling=False)


HDR_BLUE   = PatternFill('solid', fgColor='1F4E79')
HDR_MED    = PatternFill('solid', fgColor='2E75B6')
AVG_FILL   = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL   = PatternFill('solid', fgColor='FFC7CE')
AMBER_FILL = PatternFill('solid', fgColor='FFEB9C')
GRAY_FILL  = PatternFill('solid', fgColor='F2F2F2')
_thin      = Side(style='thin', color='BFBFBF')

def _bdr():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _osa_fill_xl(val):
    if val is None or (isinstance(val, float) and np.isnan(val)): return None
    return GREEN_FILL if val >= 95 else (AMBER_FILL if val >= 80 else RED_FILL)

OSA_WEEKS = ['Week 1','Week 2','Week 3','Week 4','Week 5','Week 6']
OSA_DEFAULT_PT = 6


# ─────────────────────────────────────────────────────────────────────────────
#  UPLOADED DATA PERSISTENCE  (file-backed parquet store)
#  Stores the last-uploaded OSA and SOS DataFrames so the app remembers data
#  across page navigations and browser refreshes.
# ─────────────────────────────────────────────────────────────────────────────

_DATA_STORE_DIR = pathlib.Path(tempfile.gettempdir()) / "ke_analytics_store"
_DATA_STORE_DIR.mkdir(exist_ok=True)

_OSA_STORE_PATH  = _DATA_STORE_DIR / "osa_data.pkl"
_OSA_META_PATH   = _DATA_STORE_DIR / "osa_meta.json"


def _save_df(df: pd.DataFrame, path: pathlib.Path, meta: dict, meta_path: pathlib.Path):
    import pickle
    try:
        with open(str(path), 'wb') as _f:
            pickle.dump(df, _f, protocol=4)
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, default=str))
    except Exception:
        pass


def _load_df(path: pathlib.Path, meta_path: pathlib.Path):
    """Returns (df, meta) or (None, {})."""
    import pickle
    if not path.exists():
        return None, {}
    try:
        with open(str(path), 'rb') as _f:
            df = pickle.load(_f)
        meta = json.loads(meta_path.read_text()) if meta_path.exists() else {}
        return df, meta
    except Exception:
        return None, {}


def osa_data_save(df: pd.DataFrame, filename: str = "", uploaded_by: str = ""):
    from datetime import datetime as _dt
    meta = {"filename": filename, "uploaded_by": uploaded_by,
            "saved_at": _dt.now().isoformat(), "rows": len(df)}
    _save_df(df, _OSA_STORE_PATH, meta, _OSA_META_PATH)


def osa_data_load():
    """Returns (df, meta) — df is None if no data stored."""
    return _load_df(_OSA_STORE_PATH, _OSA_META_PATH)


def osa_data_clear():
    for p in (_OSA_STORE_PATH, _OSA_META_PATH):
        if p.exists(): p.unlink()

# ─────────────────────────────────────────────────────────────────────────────
#  LARGE-FILE OPTIMISED LOADERS  (handles up to ~300 MB)
# ─────────────────────────────────────────────────────────────────────────────

# Always stream — no size threshold needed. Streaming is always safer on cloud.

def _optimise_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Downcast numeric columns and convert low-cardinality object columns to
    category dtype.  Typically cuts RAM by 50-70% on wide survey exports.
    """
    for col in df.columns:
        col_dtype = df[col].dtype
        if col_dtype == object:
            n_unique = df[col].nunique(dropna=False)
            if n_unique / max(len(df), 1) < 0.5:
                try:
                    df[col] = df[col].astype('category')
                except Exception:
                    pass
        elif col_dtype == np.float64:
            try: df[col] = df[col].astype(np.float32)
            except Exception: pass
        elif col_dtype == np.int64:
            try: df[col] = df[col].astype(np.int32)
            except Exception: pass
    return df


def read_large_csv(file_bytes: bytes, chunksize: int = 150_000) -> pd.DataFrame:
    """
    Read a large CSV as fast as possible while staying RAM-safe on ~1GB
    cloud instances.

    Tries the PyArrow engine first — a single-shot, multi-threaded, C++
    parse that is typically 3-5x faster than the row-by-row C engine on
    250-400MB SFA exports. Falls back to chunked C-engine parsing (dtype
    optimisation applied once on the full frame, not per chunk — nunique()
    per chunk was the main hot spot on the old loader).
    """
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), engine="pyarrow")
        return _optimise_dtypes(df)
    except Exception:
        pass

    chunks = []
    for chunk in pd.read_csv(
        io.BytesIO(file_bytes), chunksize=chunksize,
        low_memory=True, encoding_errors="replace"
    ):
        chunks.append(chunk)
    df = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
    return _optimise_dtypes(df)


def _bytes_to_tempfile(file_bytes: bytes, suffix: str = ".xlsx") -> str:
    """Write bytes to a named temp file and return its path.

    Using a real file path avoids keeping a second io.BytesIO copy of the
    data in RAM while openpyxl streams through it.
    """
    import tempfile as _tf
    fd, path = _tf.mkstemp(suffix=suffix)
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(file_bytes)
    except Exception:
        os.close(fd)
        raise
    return path


def read_large_excel(file_bytes: bytes, sheet_name=0) -> pd.DataFrame:
    """
    Stream-read an Excel file using openpyxl read_only mode.

    Key RAM savings vs naïve pd.read_excel():
    • Writes bytes to a temp file so openpyxl reads from disk, not a
      second in-memory BytesIO buffer.
    • Yields rows through a generator into pd.DataFrame() — never builds
      a full Python list of all rows before constructing the frame.
    • Calls wb.close() immediately after the header pass to release the
      workbook object before row iteration.

    This is safe up to ~300 MB on Streamlit Cloud's 1 GB RAM limit.
    """
    tmp = _bytes_to_tempfile(file_bytes, suffix=".xlsx")
    try:
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        # Resolve sheet
        if isinstance(sheet_name, str):
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        else:
            ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb.active

        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if header is None:
            wb.close()
            return pd.DataFrame()

        # Stream rows in chunks to avoid building a full Python list in RAM
        CHUNK = 50_000
        chunks = []
        while True:
            batch = list(itertools.islice(row_iter, CHUNK))
            if not batch:
                break
            chunks.append(_optimise_dtypes(pd.DataFrame(batch, columns=header)))
        wb.close()
        return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=header)
    finally:
        try:
            os.unlink(tmp)
        except OSError:
            pass


def smart_load(file_bytes: bytes, filename: str, sheet_name=0) -> pd.DataFrame:
    """
    Auto-detect format and load efficiently — handles CSV, XLS, XLSX up to 300 MB.
    For Excel files it first tries the named sheet ('OSA' or 'SOS'), then falls back.
    """
    fname = filename.lower()
    if fname.endswith(".csv"):
        return read_large_csv(file_bytes)
    try:
        return read_large_excel(file_bytes, sheet_name=sheet_name)
    except Exception:
        # Final fallback — treat as CSV (some SFA exports have wrong extension)
        return read_large_csv(file_bytes)


# ═════════════════════════════════════════════════════════════════════════════
#  KENYA OSA  —  Must-Have SKU (MHSKU) + MBQ pressure targets by outlet category
# ═════════════════════════════════════════════════════════════════════════════
#
# The Kenyan OSA pipeline differs from the Uganda one in three ways:
#   1. No fixed brand whitelist — the MHSKU workbook (7 sheets, one per outlet
#      category) defines which SKUs are tracked and their pressure target (MBQ),
#      per outlet category.
#   2. The SFA export sometimes carries stray Uganda/Kampala rows (cross-border
#      accounts, Export channel) — these are dropped, keeping Kenya regions only.
#   3. Pressure Target = MBQ from the sheet that matches the row's outlet
#      category — not a flat/global default.

# Outlet category (CUSTOMER CATEGORY, as it appears in the SFA OSA export) →
# matching sheet name in the Kimfay MHSKU workbook.
KE_MHSKU_SHEET_MAP = {
    'HYPER MARKET':   'Hyper',
    'LARGE SUPER':    'Large SPMKT',
    'SMALL SUPER':    'Small SPMKT',
    'EXPRESS STORES': 'Express',
}

# LMT (Local Modern Trade) rows are dropped entirely, not analysed. Two reasons:
#  1. LMT outlets aren't part of the named key-account universe (KE_KEY_ACCOUNTS) —
#     ACCOUNT NAME for these rows is the generic literal "LMT", not a named chain.
#  2. The SFA export's tier labelling for LMT is unreliable/incomplete at outlet
#     level (mostly untiered "LMT", with only a small, inconsistent number of rows
#     explicitly tagged "LMT Tier 2" / "LMT Tier 3") — guessing a tier (as an
#     earlier version of this pipeline did, mapping generic "LMT" to Tier 1)
#     produced misleading category-level OSA%.
KE_EXCLUDED_OUTLET_CATEGORIES = {
    'LMT', 'LMT TIER 1', 'LMT TIER 2', 'LMT TIER 3',
    'LMT TEIR1', 'LMT TEIR2', 'LMT TEIR3',
}

# Column name (per sheet) that holds the MBQ / pressure target — varies by sheet.
_KE_MBQ_COL_CANDIDATES = ['MBQ', 'MBQ (Small SPMKT)', 'MBQ (Express)',
                           'MBQ (LMT Teir1)', 'MBQ (LMT Teir2)', 'MBQ (LMT Teir3)']

# Kenya regions kept in the Kenyan OSA pipeline (as seen in the SFA export).
KE_REGION_KEYWORDS = ['nairobi', 'coast', 'rift valley', 'mountain', 'lake']

# Anything matching these keywords is stripped out — Uganda/Kampala leakage,
# cross-border Export-channel rows, etc.
KE_EXCLUDE_REGION_KEYWORDS = ['kampala', 'uganda', 'entebbe', 'jinja', 'mbarara',
                              'gulu', 'wakiso', 'mukono', 'lira', 'mbale', 'arua', 'masaka']

# Accounts/outlets intentionally excluded from Kenya OSA analytics — Bestmart
# chains, plus the non-tracked "MUHINDI MWEUSI WITEITHIE" outlet (not a real
# key account). Matched (case-insensitively) against CUSTOMER NAME and ACCOUNT.
KE_EXCLUDE_ACCOUNT_KEYWORDS = ['bestmart', 'best mart', 'muhindi mweusi']

# Out-of-stock reasons the SFA export records per row (populated only when
# STOCK LEVEL = "Not available"). "Not Listed" is dropped upstream (not
# tracked — see the REASON filter below); the rest explain *why* a genuine
# must-have SKU is unavailable, and drive the "why is OSA affected" insights
# panel on the OSA Analytics page.
KE_OOS_REASONS = [
    'Order Raised', 'Code Blocked', 'Out of Stock at Supply',
    'Short date at supply', 'Previous expiry history', 'Scanning Issue', 'Pricing Issue',
]

# User-friendly challenge labels and operational meaning for each raw OOS reason.
KE_OOS_REASON_GUIDANCE = {
    'Order Raised': {
        'label': 'Order not raised / late order',
        'challenge': 'Replenishment lag',
        'hint': 'The store or buyer likely did not create or submit the order in time.',
    },
    'Code Blocked': {
        'label': 'Code blocked / setup issue',
        'challenge': 'Catalogue or system setup',
        'hint': 'The SKU is blocked, not linked, or set up incorrectly in the system.',
    },
    'Out of Stock at Supply': {
        'label': 'Stock unavailable at supply',
        'challenge': 'Supply-side availability',
        'hint': 'The supplier / DC does not have the item available to fulfil demand.',
    },
    'Short date at supply': {
        'label': 'Short-dated stock at supply',
        'challenge': 'Quality / date management',
        'hint': 'Stock is too near expiry to be used and is not acceptable for sale.',
    },
    'Previous expiry history': {
        'label': 'Prior expiry / returns history',
        'challenge': 'Quality / historical expiry risk',
        'hint': 'The item has a history of expiry problems and needs tighter control.',
    },
    'Scanning Issue': {
        'label': 'Scan / inventory capture problem',
        'challenge': 'Store execution / scan compliance',
        'hint': 'The item may not have been recorded correctly in the store scan process.',
    },
    'Pricing Issue': {
        'label': 'Pricing / shelf compliance problem',
        'challenge': 'Pricing or compliance',
        'hint': 'The item is not priced or priced incorrectly, blocking the expected availability.',
    },
}

# ── Kenya key accounts — canonical name → matching keywords in CUSTOMER NAME ──
# Order matters: first match wins, so more specific keywords should sit above
# generic ones if they could ever collide (none currently do).
KE_KEY_ACCOUNTS = [
    ('QUICKMART',    ['QUICKMART', 'QUICK MART']),
    ('NAIVAS',       ['NAIVAS']),
    ('CHANDARANA',   ['CHANDARANA']),
    ('CLEANSHELF',   ['CLEANSHELF', 'CLEAN SHELF']),
    ('MAGUNAS',      ['MAGUNAS']),
    ('JAZA',         ['JAZA']),
    ('POWERSTAR',    ['POWERSTAR', 'POWER STAR']),
    ('KHETIAS',      ['KHETIAS']),
    ('CARREFOUR',    ['CARREFOUR']),
    ('LEESTAR',      ['LEESTAR']),
    ('DEFCO',        ['DEFCO']),
    ('ONN THE WAY',  ['ONN THE WAY', 'ON THE WAY']),
    ('THE ZOROS',    ['THE ZOROS', 'ZOROS']),
    ('EASTMATT',     ['EASTMATT', 'EASTMAT', 'EAST MATT']),
    ('KAMINDI',      ['KAMINDI']),
    ('KASSMATT',     ['KASSMATT', 'KASSMART']),
    ('MATHAI',       ['MATHAI']),
]

def get_ke_account(name) -> str:
    """Map a raw CUSTOMER NAME to its canonical Kenya key account, or return
    the original (stripped) name if it doesn't match any key account."""
    if pd.isna(name):
        return 'UNKNOWN'
    n = str(name).upper().strip()
    for canonical, keywords in KE_KEY_ACCOUNTS:
        if any(kw in n for kw in keywords):
            return canonical
    return str(name).strip()


# ── Shared OSA filters ────────────────────────────────────────────────────────
# The OSA Analytics page and the Report Generator use the SAME session_state
# keys so a selection made on the analytics page (e.g. only NAIROBI outlets, or
# only NAIVAS/QUICKMART) carries straight through to the downloaded report.
KE_FILTER_KEYS = {
    'months':  'ke_osa_months',
    'regions': 'ke_osa_regions',
    'ocats':   'ke_osa_ocats',
    'brands':  'ke_osa_brands',
    'accts':   'ke_osa_accts',
}


def apply_ke_osa_filters(df: pd.DataFrame, months=None, regions=None, ocats=None,
                         brands=None, accts=None) -> pd.DataFrame:
    """Apply the OSA Analytics filters to a dataframe.

    A None (or empty) selection means "no narrowing on that dimension".
    Rows with a missing REGION are always dropped (mirrors the analytics page,
    which only ever charts region-tagged Kenya rows). Returns a copy.
    """
    dff = df
    if months:
        dff = dff[dff['Month'].isin(months)]
    if 'REGION' in dff.columns:
        dff = dff[dff['REGION'].notna()]
    if regions and 'REGION' in dff.columns:
        dff = dff[dff['REGION'].isin(regions)]
    if ocats and 'MHSKU_SHEET' in dff.columns:
        dff = dff[dff['MHSKU_SHEET'].isin(ocats)]
    if brands and 'BRAND NAME' in dff.columns:
        dff = dff[dff['BRAND NAME'].isin(brands)]
    if accts and 'ACCOUNT' in dff.columns:
        dff = dff[dff['ACCOUNT'].isin(accts)]
    return dff.copy()


def load_ke_mhsku(file_bytes: bytes) -> dict:
    """
    Parse the Kimfay Kenya MHSKU workbook (Hyper / Large SPMKT / Small SPMKT /
    Express / LMT Teir1 / LMT Teir2 / LMT Teir3 sheets).

    Returns {sheet_name: {epr_code: {"mbq": int, "description": str,
                                      "brand": str, "category": str}}}
    """
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    out = {}
    for sheet in xl.sheet_names:
        if sheet.strip().lower() == 'summary':
            continue
        df = xl.parse(sheet)
        cols = {c.strip(): c for c in df.columns}
        code_col = cols.get('EPR Code')
        if not code_col:
            continue
        mbq_col = next((cols[c] for c in _KE_MBQ_COL_CANDIDATES if c in cols), None)
        if not mbq_col:
            continue
        desc_col  = cols.get('Product Description', code_col)
        brand_col = cols.get('Brand')
        cat_col   = cols.get('Category')
        sheet_map = {}
        for _, row in df.iterrows():
            code = str(row[code_col]).strip().upper()
            if not code or code == 'NAN':
                continue
            mbq_val = pd.to_numeric(row[mbq_col], errors='coerce')
            if pd.isna(mbq_val):
                continue
            sheet_map[code] = {
                "mbq": int(mbq_val),
                "description": str(row[desc_col]).strip() if desc_col else code,
                "brand": str(row[brand_col]).strip() if brand_col and pd.notna(row[brand_col]) else "",
                "category": str(row[cat_col]).strip() if cat_col and pd.notna(row[cat_col]) else "",
            }
        out[sheet.strip()] = sheet_map
    return out


# ── persistent store for the Kenya MHSKU workbook (mirrors mhsku_save/load) ──
_KE_MHSKU_STORE = pathlib.Path(tempfile.gettempdir()) / "ke_mhsku_store.json"


def ke_mhsku_save(mhsku_map: dict, uploaded_by: str = "", filename: str = ""):
    store = {
        "mhsku_map": mhsku_map,
        "last_updated": datetime.utcnow().isoformat() + "Z",
        "last_updated_by": uploaded_by,
        "filename": filename,
    }
    _KE_MHSKU_STORE.write_text(json.dumps(store, ensure_ascii=False))


def ke_mhsku_load() -> tuple[dict, str, str, str]:
    """Returns (mhsku_map, last_updated, last_updated_by, filename)."""
    if _KE_MHSKU_STORE.exists():
        try:
            store = json.loads(_KE_MHSKU_STORE.read_text())
            return (store.get("mhsku_map", {}), store.get("last_updated", ""),
                    store.get("last_updated_by", ""), store.get("filename", ""))
        except Exception:
            pass
    return {}, "", "", ""


def ke_mhsku_clear():
    if _KE_MHSKU_STORE.exists():
        _KE_MHSKU_STORE.unlink()


def _flatten_ke_mhsku(mhsku_map: dict) -> pd.DataFrame:
    """Turn {sheet: {code: {mbq,...}}} into a flat lookup frame for a vectorised merge."""
    rows = []
    for sheet, codes in (mhsku_map or {}).items():
        for code, rec in codes.items():
            rows.append((sheet, code, rec.get('mbq')))
    if not rows:
        return pd.DataFrame(columns=['MHSKU_SHEET', '_CODE_KEY', 'MBQ'])
    return pd.DataFrame(rows, columns=['MHSKU_SHEET', '_CODE_KEY', 'MBQ'])


def process_osa_kenya(file_bytes: bytes, filename: str = "",
                       mhsku_map: dict = None,
                       must_have_only: bool = True) -> pd.DataFrame:
    """
    Kenyan OSA pipeline: filters to Kenya regions, drops "Not Listed" rows
    (SKUs not tracked/ranged at that outlet), maps each row's outlet category
    (CUSTOMER CATEGORY) to the matching MHSKU sheet, and — for rows whose
    PRODUCT CODE is a must-have SKU in that sheet — sets PRESSURE TARGET to
    the sheet's MBQ and computes OSA off it.

    Performance notes (tuned for 300-400MB SFA exports):
    • CSV is read via the PyArrow engine where possible (see read_large_csv).
    • Row-count-reducing filters (region, Not Listed, must-have) all run
      BEFORE the per-row date parsing / account mapping, so the expensive
      steps only ever touch the rows that will actually appear in the report.
    • The MHSKU must-have/MBQ lookup is a vectorised pd.merge, not a
      DataFrame.apply — this alone is the difference between seconds and
      minutes on a 500k-row export.

    must_have_only=True (default): keep only must-have SKUs (recommended —
    matches how the MHSKU workbook is meant to be used as the OSA universe).
    must_have_only=False: keep all rows, tagging MUST HAVE True/False.
    """
    fname = filename.lower()
    if fname.endswith(".csv"):
        df = read_large_csv(file_bytes)
    else:
        try:
            df = read_large_excel(file_bytes, sheet_name=0)
        except Exception:
            df = read_large_csv(file_bytes)

    # ── normalise column names (reuse OSA aliasing) ──────────────────────────
    _OSA_DATE_ALIASES = [
        "date reported", "date_reported", "visit date", "date created",
        "date_created", "survey date", "submission date", "date", "created date",
    ]
    col_lower = {c.lower().strip(): c for c in df.columns}
    if 'DATE REPORTED' not in df.columns:
        for alias in _OSA_DATE_ALIASES:
            if alias in col_lower:
                df = df.rename(columns={col_lower[alias]: 'DATE REPORTED'})
                break
        else:
            available = ", ".join(f"'{c}'" for c in df.columns[:20])
            raise ValueError(
                f"Could not find a date column in your OSA file. "
                f"Expected 'DATE REPORTED' or similar. Columns found: {available}"
            )

    _OSA_COL_ALIASES = {
        "CUSTOMER NAME":     ["customer name", "customer_name", "outlet", "outlet name", "store name", "store"],
        "BRAND NAME":        ["brand name", "brand_name", "brand", "product brand"],
        "PRODUCT CATEGORY":  ["product category", "product_category", "category", "cat"],
        "PRODUCT CODE":      ["product code", "product_code", "sku code", "item code", "code"],
        "DESCRIPTION":       ["description", "product description", "product_description",
                               "product name", "product_name", "sku name", "sku_name",
                               "item description", "item name", "item_name", "sku description"],
        "PRESSURE TARGET":   ["pressure target", "pressure_target", "target", "pt"],
        "QUANTITY":          ["quantity", "qty", "stock qty", "count"],
        "STOCK LEVEL":       ["stock level", "stock_level", "availability", "available"],
        "REGION":            ["region", "territory", "area", "district"],
        "CUSTOMER CATEGORY": ["customer category", "customer_category", "outlet category", "outlet type"],
        "REASON":            ["reason", "out of stock reason", "oos reason"],
    }
    col_lower = {c.lower().strip(): c for c in df.columns}
    renames = {}
    for canonical, aliases in _OSA_COL_ALIASES.items():
        if canonical in df.columns:
            continue
        for alias in aliases:
            if alias in col_lower:
                renames[col_lower[alias]] = canonical
                break
    if renames:
        df = df.rename(columns=renames)

    # ── keep Kenya regions only, drop Kampala/Uganda leakage (vectorised) ────
    if 'REGION' in df.columns:
        region_norm = df['REGION'].astype(str).str.strip().str.lower()
        pattern = '|'.join(KE_EXCLUDE_REGION_KEYWORDS)
        df = df[~region_norm.str.contains(pattern, regex=True, na=False)].copy()

    # ── drop Bestmart accounts from the Kenya key-account scope ──────────────
    acct_pattern = '|'.join(KE_EXCLUDE_ACCOUNT_KEYWORDS)
    if 'CUSTOMER NAME' in df.columns:
        customer_norm = df['CUSTOMER NAME'].astype(str).str.strip().str.lower()
        df = df[~customer_norm.str.contains(acct_pattern, regex=True, na=False)].copy()
    if 'ACCOUNT' in df.columns:
        acct_norm = df['ACCOUNT'].astype(str).str.strip().str.lower()
        df = df[~acct_norm.str.contains(acct_pattern, regex=True, na=False)].copy()

    # ── drop "Not Listed" rows — SKUs not ranged at that outlet aren't tracked
    if 'REASON' in df.columns:
        reason_norm = df['REASON'].astype(str).str.strip().str.lower()
        df = df[reason_norm != 'not listed'].copy()

    # ── drop LMT (all tiers) — not a named key account, tier data unreliable ──
    if 'CUSTOMER CATEGORY' in df.columns:
        cat_norm = df['CUSTOMER CATEGORY'].astype(str).str.strip().str.upper()
        df = df[~cat_norm.isin(KE_EXCLUDED_OUTLET_CATEGORIES)].copy()

    # ── MHSKU must-have + MBQ mapping (vectorised merge, not row-wise apply) ─
    mhsku_map = mhsku_map or {}
    df['MHSKU_SHEET'] = (df.get('CUSTOMER CATEGORY', pd.Series(index=df.index, dtype=object))
                            .astype(str).str.strip().str.upper()
                            .map(KE_MHSKU_SHEET_MAP))

    if 'PRODUCT CODE' in df.columns and mhsku_map:
        df['_CODE_KEY'] = df['PRODUCT CODE'].astype(str).str.strip().str.upper()
        lookup = _flatten_ke_mhsku(mhsku_map)
        df = df.merge(lookup, on=['MHSKU_SHEET', '_CODE_KEY'], how='left')
        df['MUST HAVE'] = df['MBQ'].notna()
        df['PRESSURE TARGET'] = df['MBQ']
        df = df.drop(columns=['_CODE_KEY', 'MBQ'])
    else:
        df['MUST HAVE'] = False
        df['PRESSURE TARGET'] = pd.to_numeric(df.get('PRESSURE TARGET', 0), errors='coerce')

    if must_have_only and mhsku_map:
        df = df[df['MUST HAVE']].copy()

    if df.empty:
        return df

    # ── parse dates (now running only on the rows that survive filtering) ───
    if pd.api.types.is_datetime64_any_dtype(df['DATE REPORTED']):
        # Already parsed (e.g. the PyArrow CSV engine auto-detects date columns) —
        # use as-is, this is the fast path and avoids double-interpreting it as
        # an Excel serial number.
        df['DATE_PARSED'] = df['DATE REPORTED']
        _date_col = df['DATE REPORTED'].dropna()
        _is_numeric = False
    else:
        _date_col = df['DATE REPORTED'].dropna()
        _is_numeric = pd.to_numeric(_date_col, errors='coerce').notna().mean() > 0.8
    if _is_numeric:
        _serials = pd.to_numeric(df['DATE REPORTED'], errors='coerce')
        df['DATE_PARSED'] = pd.Timestamp('1899-12-30') + pd.to_timedelta(_serials, unit='D')
    elif 'DATE_PARSED' not in df.columns:
        _sample = _date_col.astype(str).head(20)
        _is_iso = _sample.str.match(r'^\d{4}[-/]\d').any()
        df['DATE_PARSED'] = pd.to_datetime(df['DATE REPORTED'], errors='coerce', dayfirst=not _is_iso)
    df = df[df['DATE_PARSED'].notna()].copy()

    if 'Month' not in df.columns:
        df['Month'] = df['DATE_PARSED'].dt.strftime('%B')
    df['WEEK_NUM']   = df['DATE_PARSED'].dt.isocalendar().week.astype('Int64')
    df['WEEK_LABEL'] = None
    for month, group in df.groupby('Month', observed=True):
        sw = sorted(group['WEEK_NUM'].dropna().unique())
        wm = {w: f'Week {i+1}' for i, w in enumerate(sw)}
        df.loc[df['Month'] == month, 'WEEK_LABEL'] = df.loc[df['Month'] == month, 'WEEK_NUM'].map(wm)

    df['QUANTITY'] = pd.to_numeric(df.get('QUANTITY', 0), errors='coerce').fillna(0)

    # ── key accounts ─────────────────────────────────────────────────────
    if 'CUSTOMER NAME' in df.columns:
        df['ACCOUNT'] = df['CUSTOMER NAME'].apply(get_ke_account)

    df['PRESSURE TARGET'] = pd.to_numeric(df['PRESSURE TARGET'], errors='coerce').fillna(OSA_DEFAULT_PT)
    df['OSA'] = (df['QUANTITY'] >= df['PRESSURE TARGET']).astype(int) * 100

    return df


# ─────────────────────────────────────────────────────────────────────────────
#  KENYA — formatted Excel report (Report Generator page)
# ─────────────────────────────────────────────────────────────────────────────

def build_osa_excel_kenya(df: pd.DataFrame) -> bytes:
    """Build the formatted Kenya OSA Excel workbook — mirrors the Uganda
    report layout (one tab per month x breakdown, colour-coded, legend tab)
    but grouped by Outlet Category / Account / SKU instead of Uganda brands."""

    def make_pivot(data, rows):
        wp = [w for w in OSA_WEEKS if w in data['WEEK_LABEL'].values]
        pt = data.pivot_table(index=rows, columns='WEEK_LABEL', values='OSA', aggfunc='mean')
        pt = pt.reindex(columns=[w for w in wp if w in pt.columns])
        pt['Average'] = pt.mean(axis=1)
        return pt.round(1)

    def write_sheet(ws, pivot_df, title, row_labels):
        ws.freeze_panes = f'{get_column_letter(len(row_labels)+1)}3'
        all_cols   = list(pivot_df.columns)
        n_rc       = len(row_labels)
        total_cols = n_rc + len(all_cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        tc = ws.cell(1, 1, title)
        tc.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        tc.fill = HDR_BLUE; tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22
        for ci, lbl in enumerate(row_labels, 1):
            c = ws.cell(2, ci, lbl)
            c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
            c.fill = HDR_MED; c.alignment = Alignment(horizontal='center', vertical='center'); c.border = _bdr()
        for ci, col in enumerate(all_cols, n_rc+1):
            is_avg = col == 'Average'
            c = ws.cell(2, ci, col)
            c.font = Font(name='Calibri', bold=True, color='7F3F00' if is_avg else 'FFFFFF', size=10)
            c.fill = AVG_FILL if is_avg else HDR_MED
            c.alignment = Alignment(horizontal='center'); c.border = _bdr()
        ws.row_dimensions[2].height = 18
        pr = pivot_df.reset_index(); pf = None
        for ri, row in pr.iterrows():
            er = ri+3; bg = GRAY_FILL if ri%2==0 else None
            fv = row[row_labels[0]] if row_labels[0] in row.index else row.iloc[0]
            is_new = fv != pf
            # collapsible outline: tuck continuation rows under each first-level value
            if n_rc > 1 and not is_new:
                ws.row_dimensions[er].outline_level = 1
            for ci, lbl in enumerate(row_labels, 1):
                val = row[lbl] if lbl in row.index else row.iloc[ci-1]
                disp = (str(val) if pd.notna(val) else '') if (ci>1 or is_new) else ''
                c = ws.cell(er, ci, disp)
                c.font = Font(name='Calibri', size=10, bold=(ci==1))
                c.fill = bg or PatternFill()
                c.alignment = Alignment(vertical='center', horizontal='left', indent=1); c.border = _bdr()
            for ci, col in enumerate(all_cols, n_rc+1):
                val = row.get(col, np.nan)
                if pd.isna(val):
                    c = ws.cell(er, ci, '-')
                    c.font = Font(name='Calibri', size=10, color='BFBFBF'); c.fill = bg or PatternFill()
                else:
                    c = ws.cell(er, ci, round(val,1)/100); c.number_format = '0%'
                    f = _osa_fill_xl(val)
                    c.fill = f if f else (AVG_FILL if col=='Average' else (bg or PatternFill()))
                    c.font = Font(name='Calibri', size=10, bold=(col=='Average'))
                c.alignment = Alignment(horizontal='center', vertical='center'); c.border = _bdr()
            pf = fv
        # filter dropdowns on the header row + collapse groups from the top row
        last_row = len(pr) + 2
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{last_row}"
        if n_rc > 1:
            ws.sheet_properties.outlinePr.summaryBelow = False
        for i in range(n_rc):
            lbl = row_labels[i]
            default_w = 40 if lbl == 'SKU' else (28 if i == 0 else 22)
            ws.column_dimensions[get_column_letter(i+1)].width = default_w
        for i in range(len(all_cols)):
            ws.column_dimensions[get_column_letter(n_rc+i+1)].width = 10

    wb = openpyxl.Workbook(); first = True
    months = sorted(df['Month'].dropna().unique(), key=lambda m: datetime.strptime(m, '%B').month)

    def _safe_sheet_name(label: str) -> str:
        s = re.sub(r'[^A-Za-z0-9 _\-()]', '_', str(label))
        return s[:31]

    df = df.copy()
    col_map = {c.upper().strip(): c for c in df.columns}
    desc_candidates = [
        'DESCRIPTION', 'PRODUCT DESCRIPTION', 'PRODUCT_DESCRIPTION',
        'PRODUCT NAME', 'PRODUCT_NAME', 'SKU', 'SKU NAME', 'SKU_NAME',
        'ITEM DESCRIPTION', 'ITEM_DESCRIPTION'
    ]
    desc_col = None
    for cand in desc_candidates:
        if cand in col_map:
            desc_col = col_map[cand]
            break
    if desc_col:
        df['SKU'] = df[desc_col].astype(str).str.strip()
    elif 'PRODUCT CODE' in df.columns:
        df['SKU'] = df['PRODUCT CODE'].astype(str).str.strip()

    configs = [
        ('Outlet Cat OSA',      ['MHSKU_SHEET'],           'OSA — BY OUTLET CATEGORY (MBQ TARGET)'),
        ('Account OSA',         ['ACCOUNT'],               'OSA — BY KEY ACCOUNT'),
        ('Acct x Outlet Cat',   ['ACCOUNT','MHSKU_SHEET'], 'OSA — BY ACCOUNT & OUTLET CATEGORY'),
    ]
    # Real product category (PRODUCT CATEGORY = MOUSSE, LIQUID SOAP, …) — this is
    # distinct from the outlet category (MHSKU_SHEET = Hyper / Large SPMKT / …).
    if 'PRODUCT CATEGORY' in df.columns:
        configs += [
            ('Product Cat OSA',   ['PRODUCT CATEGORY'],           'OSA — BY PRODUCT CATEGORY'),
            ('Acct x Product Cat', ['ACCOUNT','PRODUCT CATEGORY'], 'OSA — BY ACCOUNT & PRODUCT CATEGORY'),
        ]
    if 'BRAND NAME' in df.columns:
        configs.append(('Account x Brand', ['ACCOUNT','BRAND NAME'], 'OSA — BY ACCOUNT & BRAND'))
    if 'SKU' in df.columns:
        configs += [
            ('SKU OSA',       ['SKU'],            'OSA — BY MUST-HAVE SKU'),
            ('Account x SKU', ['ACCOUNT','SKU'],  'OSA — BY ACCOUNT & SKU'),
        ]

    ms = []
    for month in months:
        md = df[df['Month'] == month]
        wk = sorted(md['WEEK_LABEL'].dropna().unique(), key=lambda w: int(w.split()[-1]))
        ms.append((month, f"{len(md):,} rows | {len(wk)} week(s): {', '.join(wk)}"))
        for tb, rl, tt in configs:
            if not all(c in md.columns for c in rl):
                continue
            tn = f"{month[:3]} - {tb}"
            ws = wb.active if first else wb.create_sheet(tn)
            if first: ws.title = tn; first = False
            write_sheet(ws, make_pivot(md, rl), f"{tt} — {month.upper()} (Kenya)", rl)

    # ── OOS Reason Analysis — why OSA is affected ────────────────────────────
    if 'REASON' in df.columns:
        gap = df[df['OSA'] == 0].copy()
        gap['REASON'] = gap['REASON'].astype(str).str.strip()
        gap = gap[gap['REASON'].isin(KE_OOS_REASONS)]
        gap = gap[gap['ACCOUNT'].notna()].copy() if 'ACCOUNT' in gap.columns else gap
        if not gap.empty:
            # One sheet per account that highlights the top 5 affected branches and the dominant reason.
            acct_rank = (gap.groupby('ACCOUNT').size().reset_index(name='Total Gaps')
                         .sort_values('Total Gaps', ascending=False)
                         .reset_index(drop=True))
            has_region = 'REGION' in gap.columns
            has_rep    = 'REP NAME' in gap.columns
            has_super  = 'SUPERVISOR' in gap.columns
            has_cat    = 'PRODUCT CATEGORY' in gap.columns
            has_brand  = 'BRAND NAME' in gap.columns
            has_sku    = 'SKU' in gap.columns
            has_code   = 'PRODUCT CODE' in gap.columns

            # Combined SKU label (code — description) for the affected-SKU lists.
            if has_code and has_sku:
                gap['_SKU_LABEL'] = (gap['PRODUCT CODE'].astype(str).str.strip()
                                     + ' — ' + gap['SKU'].astype(str).str.strip())
            elif has_sku:
                gap['_SKU_LABEL'] = gap['SKU'].astype(str).str.strip()
            elif has_code:
                gap['_SKU_LABEL'] = gap['PRODUCT CODE'].astype(str).str.strip()
            has_skulabel = '_SKU_LABEL' in gap.columns

            def _mode_or_join(series, limit=3):
                vals = [str(v) for v in series.dropna().unique()]
                return ', '.join(vals[:limit]) + (' …' if len(vals) > limit else '')

            def _write_mini_table(ws, start, title, counts, label_hdr, top=10):
                """Write a compact 2-column 'label | Gap Rows' table (cols A:B).
                Returns the next free row (a blank spacer already left)."""
                ws.cell(start, 1, title).font = Font(name='Calibri', bold=True, size=10, color='1F4E78')
                hr = start + 1
                for ci, h in enumerate((label_hdr, 'Gap Rows'), 1):
                    c = ws.cell(hr, ci, h)
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF'); c.fill = HDR_MED; c.border = _bdr()
                last = hr
                for i, (lbl, cnt) in enumerate(counts.head(top).items()):
                    last = hr + 1 + i
                    fill = GRAY_FILL if i % 2 == 0 else PatternFill()
                    a = ws.cell(last, 1, str(lbl)); a.font = Font(name='Calibri', size=10); a.fill = fill; a.border = _bdr()
                    b = ws.cell(last, 2, int(cnt)); b.font = Font(name='Calibri', size=10); b.fill = fill; b.border = _bdr()
                return last + 2

            for acct in acct_rank['ACCOUNT']:
                acct_gap = gap[gap['ACCOUNT'] == acct].copy()
                acct_name = _safe_sheet_name(f"Account - {acct}")
                ws_acc = wb.create_sheet(acct_name)
                ws_acc.merge_cells('A1:H1')
                ws_acc['A1'] = f'ACCOUNT ROOT CAUSE DETAIL — {acct}'
                ws_acc['A1'].font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
                ws_acc['A1'].fill = HDR_BLUE
                ws_acc['A1'].alignment = Alignment(horizontal='center', vertical='center')

                # ── summary strip: total gaps, regions, and who to follow up with ──
                summary = [('Account', str(acct)), ('Total gaps', int(len(acct_gap)))]
                if has_region:
                    summary.append(('Region(s)', _mode_or_join(acct_gap['REGION'], limit=6)))
                if has_rep:
                    summary.append(('Field reps to follow up', _mode_or_join(acct_gap['REP NAME'], limit=8)))
                if has_super:
                    summary.append(('Supervisor(s)', _mode_or_join(acct_gap['SUPERVISOR'], limit=6)))
                for i, (k, v) in enumerate(summary):
                    r = 2 + i
                    ws_acc.cell(r, 1, k).font = Font(name='Calibri', bold=True, color='FFFFFF')
                    ws_acc.cell(r, 1).fill = HDR_MED
                    ws_acc.cell(r, 1).border = _bdr()
                    ws_acc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
                    vc = ws_acc.cell(r, 2, v)
                    vc.font = Font(name='Calibri', bold=(i == 0))
                    vc.fill = GRAY_FILL
                    vc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

                # ── Top branches: region · dominant reason · who filled (follow-up) ──
                branch_rows = []
                for br, br_df in acct_gap.groupby('CUSTOMER NAME', observed=True):
                    br_reason = br_df.groupby('REASON', observed=True).size().sort_values(ascending=False)
                    top_reason = str(br_reason.index[0]) if len(br_reason) else ''
                    top_reason_rows = int(br_reason.iloc[0]) if len(br_reason) else 0
                    br_region = (str(br_df['REGION'].dropna().mode().iloc[0])
                                 if has_region and br_df['REGION'].notna().any() else '')
                    if has_rep and br_df['REP NAME'].notna().any():
                        rep_counts = br_df.groupby('REP NAME', observed=True).size().sort_values(ascending=False)
                        top_rep, top_rep_rows = str(rep_counts.index[0]), int(rep_counts.iloc[0])
                    else:
                        top_rep, top_rep_rows = '', 0
                    branch_rows.append([str(br), br_region, int(len(br_df)),
                                        top_reason, top_reason_rows, top_rep, top_rep_rows])
                branch_tbl = sorted(branch_rows, key=lambda x: x[2], reverse=True)[:8]

                sec_row = 2 + len(summary) + 1
                ws_acc.cell(sec_row, 1, 'TOP BRANCHES — REGION · DOMINANT REASON · WHO TO FOLLOW UP').font = \
                    Font(name='Calibri', bold=True, size=10, color='1F4E78')
                hdr_row = sec_row + 1
                b_headers = ['Branch', 'Region', 'Total Gaps', 'Top Reason',
                             'Top Reason Rows', 'Filled By (Rep)', 'Rep Gap Rows']
                for ci, hdr in enumerate(b_headers, 1):
                    c = ws_acc.cell(hdr_row, ci, hdr)
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF')
                    c.fill = HDR_MED
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = _bdr()
                for ri, vals in enumerate(branch_tbl):
                    er = hdr_row + 1 + ri
                    for ci, val in enumerate(vals, 1):
                        c = ws_acc.cell(er, ci, val)
                        c.font = Font(name='Calibri', size=10)
                        c.fill = GRAY_FILL if ri % 2 == 0 else PatternFill()
                        c.border = _bdr()

                # ── Account-level affected categories / brands / SKUs (cols A:B) ──
                cur = hdr_row + len(branch_tbl) + 2
                if has_cat:
                    cur = _write_mini_table(
                        ws_acc, cur, 'AFFECTED PRODUCT CATEGORIES (account)',
                        acct_gap.groupby('PRODUCT CATEGORY', observed=True).size().sort_values(ascending=False),
                        'Product Category', top=10)
                if has_brand:
                    cur = _write_mini_table(
                        ws_acc, cur, 'AFFECTED BRANDS (account)',
                        acct_gap.groupby('BRAND NAME', observed=True).size().sort_values(ascending=False),
                        'Brand', top=10)
                if has_skulabel:
                    cur = _write_mini_table(
                        ws_acc, cur, 'AFFECTED SKUs (account)',
                        acct_gap.groupby('_SKU_LABEL', observed=True).size().sort_values(ascending=False),
                        'SKU', top=15)

                # ── Branch-level detail — category, brand & SKU affected, with reason,
                #    region and rep for direct follow-up (full SKU-level worklist) ─────
                key_label = {'CUSTOMER NAME': 'Branch', 'REGION': 'Region',
                             'PRODUCT CATEGORY': 'Category', 'BRAND NAME': 'Brand',
                             '_SKU_LABEL': 'SKU', 'REASON': 'Reason', 'REP NAME': 'Filled By (Rep)'}
                key_width = {'CUSTOMER NAME': 42, 'REGION': 14, 'PRODUCT CATEGORY': 18,
                             'BRAND NAME': 22, '_SKU_LABEL': 44, 'REASON': 20, 'REP NAME': 22}
                detail_keys = ['CUSTOMER NAME']
                if has_region:   detail_keys.append('REGION')
                if has_cat:      detail_keys.append('PRODUCT CATEGORY')
                if has_brand:    detail_keys.append('BRAND NAME')
                if has_skulabel: detail_keys.append('_SKU_LABEL')
                detail_keys.append('REASON')
                if has_rep:      detail_keys.append('REP NAME')
                detail = (acct_gap.groupby(detail_keys, observed=True).size()
                          .reset_index(name='Rows')
                          .sort_values(['CUSTOMER NAME', 'Rows'], ascending=[True, False])
                          .head(3000))
                d_headers = [key_label[k] for k in detail_keys] + ['Rows']

                start_row = cur
                ws_acc.cell(start_row - 1, 1, 'BRANCH × CATEGORY × BRAND × SKU DETAIL '
                            '(affected SKUs per branch — reason, region & rep for follow-up)').font = \
                    Font(name='Calibri', bold=True, size=10, color='1F4E78')
                for ci, hdr in enumerate(d_headers, 1):
                    c = ws_acc.cell(start_row, ci, hdr)
                    c.font = Font(name='Calibri', bold=True, color='FFFFFF')
                    c.fill = HDR_MED
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = _bdr()
                n_keys = len(detail_keys)
                detail_r = detail.reset_index(drop=True)
                prev_branch = None
                for ri, (_, row) in enumerate(detail_r.iterrows()):
                    er = start_row + 1 + ri
                    fill = GRAY_FILL if ri % 2 == 0 else PatternFill()
                    # collapsible outline: group each branch's SKU rows under its first row
                    branch_val = row['CUSTOMER NAME']
                    if branch_val == prev_branch:
                        ws_acc.row_dimensions[er].outline_level = 1
                    prev_branch = branch_val
                    for ci, key in enumerate(detail_keys, 1):
                        val = row[key]
                        c = ws_acc.cell(er, ci, str(val) if pd.notna(val) else '')
                        c.font = Font(name='Calibri', size=10); c.fill = fill; c.border = _bdr()
                    c = ws_acc.cell(er, n_keys + 1, int(row['Rows']))
                    c.font = Font(name='Calibri', size=10); c.fill = fill; c.border = _bdr()
                    c.alignment = Alignment(horizontal='center')

                # filter dropdowns over the detail header + rows, collapse from top row
                last_detail_row = start_row + len(detail_r)
                ws_acc.auto_filter.ref = f"A{start_row}:{get_column_letter(n_keys + 1)}{last_detail_row}"
                ws_acc.sheet_properties.outlinePr.summaryBelow = False

                for i, k in enumerate(detail_keys):
                    ws_acc.column_dimensions[get_column_letter(i + 1)].width = key_width.get(k, 16)
                ws_acc.column_dimensions[get_column_letter(n_keys + 1)].width = 10
                for r in range(1, sec_row):
                    ws_acc.row_dimensions[r].height = 16

            ws_r = wb.create_sheet('OOS Reason Analysis')
            group_col = 'ACCOUNT' if 'ACCOUNT' in gap.columns else 'MHSKU_SHEET'
            branches = gap.groupby(group_col)['MHSKU_SHEET'].agg(
                lambda s: ', '.join(sorted(map(str, s.dropna().unique())))
            )
            pt = (gap.groupby([group_col, 'REASON']).size().unstack(fill_value=0))
            pt['Total'] = pt.sum(axis=1)
            pt = pt.sort_values('Total', ascending=False)
            pt = pt.join(branches.rename('Branches'))
            pt['Account (Branches)'] = pt.apply(
                lambda r: f"{r.name} ({r['Branches']})" if pd.notna(r['Branches']) and str(r['Branches']).strip() else str(r.name),
                axis=1,
            )
            all_cols = [c for c in pt.columns if c not in {'Branches', 'Account (Branches)', 'Total'}]
            all_cols.append('Total')
            n_rc = 1
            total_cols = n_rc + len(all_cols)
            ws_r.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            tc = ws_r.cell(1, 1, 'WHY OSA IS AFFECTED — OOS REASON BY ACCOUNT (Kenya)')
            tc.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
            tc.fill = HDR_BLUE; tc.alignment = Alignment(horizontal='center', vertical='center')
            ws_r.cell(2, 1, 'Account (Branches)').font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
            ws_r.cell(2, 1).fill = HDR_MED; ws_r.cell(2, 1).border = _bdr()
            for ci, col in enumerate(all_cols, 2):
                c = ws_r.cell(2, ci, col)
                is_total = col == 'Total'
                c.font = Font(name='Calibri', bold=True, color='7F3F00' if is_total else 'FFFFFF', size=10)
                c.fill = AVG_FILL if is_total else HDR_MED
                c.alignment = Alignment(horizontal='center'); c.border = _bdr()
            pr = pt.reset_index()
            for ri, row in pr.iterrows():
                er = ri + 3; bg = GRAY_FILL if ri % 2 == 0 else None
                c = ws_r.cell(er, 1, str(row['Account (Branches)']))
                c.font = Font(name='Calibri', size=10, bold=True); c.fill = bg or PatternFill()
                c.alignment = Alignment(vertical='center', horizontal='left', indent=1); c.border = _bdr()
                for ci, col in enumerate(all_cols, 2):
                    val = row[col]
                    c = ws_r.cell(er, ci, int(val))
                    c.font = Font(name='Calibri', size=10, bold=(col == 'Total'))
                    c.fill = AVG_FILL if col == 'Total' else (bg or PatternFill())
                    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = _bdr()
            ws_r.column_dimensions['A'].width = 40
            for i in range(len(all_cols)):
                ws_r.column_dimensions[get_column_letter(i + 2)].width = 18
            ws_r.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{len(pr) + 2}"

    wl = wb.create_sheet('Legend')
    wl.merge_cells('A1:D1'); wl['A1'] = 'KENYA OSA REPORT — LEGEND & METHODOLOGY'
    wl['A1'].font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    wl['A1'].fill = HDR_BLUE; wl['A1'].alignment = Alignment(horizontal='center')
    notes = [('',''),('COLOR CODING',''),
        ('Green  ≥ 95%','On-shelf availability at or above target'),
        ('Amber  80–94%','Near target – monitor closely'),
        ('Red  < 80%','Below target – requires attention'),
        ('',''),('METHODOLOGY',''),
        ('Step 1','Kenya regions only — Kampala/Uganda rows dropped.'),
        ('Step 2','"Not Listed" rows (SKU not ranged at that outlet) dropped — not tracked.'),
        ('Step 3','LMT (all tiers) dropped — not a named key account and tier data is unreliable in the SFA export.'),
        ('Step 4','Only Must-Have SKUs (per the MHSKU workbook) for each outlet\'s category are scored.'),
        ('Step 5','Pressure Target = MBQ from the MHSKU sheet matching the outlet\'s category '
                   '(Hyper, Large SPMKT, Small SPMKT, Express).'),
        ('Step 6','Quantity ≥ Pressure Target (MBQ) → 100%  |  Quantity < Pressure Target → 0%'),
        ('Step 7','Weeks numbered 1,2,3… in order within each month.'),
        ('Step 8','Mean OSA score per Outlet Category / Account / SKU grouping.'),
        ('Step 9','"OOS Reason Analysis" sheet explains WHY must-have SKUs below target are unavailable '
                   '(Order Raised, Code Blocked, Out of Stock at Supply, etc.) — see that tab.'),
        ('',''),('KEY ACCOUNTS TRACKED',''),
        ('', ', '.join(c for c, _ in KE_KEY_ACCOUNTS)),
        ('',''),('DATA SUMMARY',''),
    ] + [(f'  {m}', s) for m, s in ms]
    for ri, (k, v) in enumerate(notes, 3):
        wl[f'A{ri}'] = k; wl[f'B{ri}'] = v
        wl[f'A{ri}'].font = Font(name='Calibri', bold=any(x in k for x in ('CODING','METHODOLOGY','DATA','Step','KEY')), size=10)
        wl[f'B{ri}'].font = Font(name='Calibri', size=10)
    wl.column_dimensions['A'].width = 32; wl.column_dimensions['B'].width = 90

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
